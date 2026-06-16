#!/usr/bin/env python3
"""
gen-sku-vials-edit.py — Two-stage vial generator that matches the real
Merit print-handoff label spec:

  STEP A — Generate ONE canonical base vial: photoreal cobalt-capped vial
           with the actual Merit label (cobalt 'RESEARCH USE ONLY' pill,
           giant grey 'M.' watermark, cobalt dot) but a BLANK area on the
           left for product text.

  STEP B — For each SKU, call gpt-image-1's EDIT endpoint with that base
           image + a prompt that adds only the compound name + size into
           the blank area. The vial, label background, RUO pill, M
           watermark, and cap all stay identical SKU-to-SKU.

USAGE
-----
  # Render the base vial only (~$0.04 medium / $0.19 high)
  python3 scripts/gen-sku-vials-edit.py --base

  # Render the base + ONE example SKU edit (for review)
  python3 scripts/gen-sku-vials-edit.py --example retatrutide-30mg

  # Render the full batch (~66 edits, ~$2.64 medium)
  python3 scripts/gen-sku-vials-edit.py --all

  # Higher fidelity (~$12.73 high quality for 66)
  python3 scripts/gen-sku-vials-edit.py --all --quality high

  # Re-edit a subset if some came out wrong
  python3 scripts/gen-sku-vials-edit.py --only retatrutide,tirzepatide

OUTPUT
------
  public/products/_base-vial.png            — canonical base, one shot
  public/products/sku-<slug>.webp           — one per SKU, edited from base
  public/products/sku-image-manifest.json   — slug → imageUrl map (read
                                              by the inventory importer)

ENV
---
  OPENAI_API_KEY — pulled from ../Merit Peptides/.env
"""

import os
import sys
import re
import json
import time
import base64
import subprocess
from pathlib import Path

HERE = Path(__file__).parent.parent
REPO_ROOT = HERE
PEPTIDES_ENV = REPO_ROOT.parent / "Merit Peptides" / ".env"
INVENTORY = REPO_ROOT.parent / "Merit Peptides" / "Inventory 6.14.xlsx"
OUTPUT_DIR = REPO_ROOT / "public" / "products"
BASE_PATH = OUTPUT_DIR / "_base-vial.png"
MASK_PATH = OUTPUT_DIR / "_edit-mask.png"
MANIFEST_PATH = OUTPUT_DIR / "sku-image-manifest.json"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Load OPENAI_API_KEY ────────────────────────────────────────────────────
for env_path in (PEPTIDES_ENV, REPO_ROOT / ".env.local"):
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if "=" in line and not line.strip().startswith("#"):
                k, _, v = line.partition("=")
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

# ── Prompts ────────────────────────────────────────────────────────────────
# Base vial prompt — matches the real Merit print-handoff label spec from
# /Merit-Labels-Print-Handoff/_PROOF-actual-size.png:
#   • cream/off-white landscape label wrapped around lower half of vial
#   • upper-left of label: tiny cobalt blue "RESEARCH USE ONLY" pill badge
#   • center-right of label: LARGE pale-grey "M." watermark monogram
#   • right of the M, vertically centered: small solid cobalt blue dot
#   • LEFT side of label: BLANK area (text gets edited in per-SKU)
def per_sku_prompt(compound: str, vial_size: str) -> str:
    """Single-shot prompt that bakes the full Merit label spec into the
    AI render — no flat overlay. Designed for gpt-image-1.generate at
    high quality. Mirrors the print-handoff proof spec exactly."""
    return (
        f"Photorealistic studio product photograph of a single "
        f"pharmaceutical peptide vial, centered, square 1:1 framing. "
        f"The aluminum crimp cap is deep cobalt blue (#2E4DDB) with a "
        f"slight metallic sheen. Clear glass body. Lyophilized white "
        f"powder fills the bottom one third of the vial. "
        f"Wrapped around the lower half of the vial is a clean white "
        f"landscape rectangular label with FOUR elements arranged as "
        f"follows: "
        f"(1) UPPER-LEFT of the label: a small cobalt blue pill-shaped "
        f"badge reading 'RESEARCH USE ONLY' in tiny white all-caps sans-"
        f"serif text. "
        f"(2) LEFT-CENTER of the label, directly below the badge: the "
        f"compound name '{compound}' in BOLD modern sans-serif (Helvetica "
        f"Bold or Inter Bold), black ink, large size. "
        f"(3) Directly below the compound name, tighter line spacing: "
        f"the vial size '{vial_size}' in a lighter weight, smaller font, "
        f"muted dark grey. "
        f"(4) RIGHT side of the label: a LARGE pale beige-grey letter 'M' "
        f"followed by a period rendered in MODERN GEOMETRIC SANS-SERIF "
        f"(Inter Black or Helvetica Black — clean uniform strokes, NO "
        f"serifs, NOT Times or Bodoni). The M is roughly the full height "
        f"of the label, soft beige-grey #D8D5CD, occupying the right "
        f"30% of the label as a watermark. Immediately to the right of "
        f"the M's period, vertically centered, a small solid cobalt blue "
        f"circle (brand dot). "
        f"Soft warm studio lighting, gentle shadow under the vial, "
        f"shallow depth of field. Background is a smooth warm cream "
        f"off-white seamless studio sweep (#F4F1EA). No props, no "
        f"documents, no humans, no hands, no extra text anywhere. "
        f"Editorial pharma-brand quality, Apple product photography. "
        f"Absolutely photoreal — not an illustration, not a 3D render."
    )


BASE_PROMPT = (
    "Photorealistic studio product photograph of a single pharmaceutical "
    "peptide vial — DEPRECATED, kept for the --base mode. Use "
    "per_sku_prompt() for per-SKU single-shot renders."
)


def edit_prompt(compound: str, vial_size: str) -> str:
    # When using a mask, only the transparent area of the mask gets
    # edited. Everything else is preserved pixel-identical — so this
    # prompt only describes what to PUT in the masked region.
    return (
        f"Fill the empty white label region with two lines of left-aligned "
        f"product text on a clean white background matching the surrounding "
        f"label: "
        f"Line 1 (top): the compound name '{compound}' in MEDIUM-BOLD "
        f"modern sans-serif (Inter Semibold or Helvetica Bold), black ink "
        f"color. If the name is long, use a smaller font size so it fits "
        f"comfortably within the region. "
        f"Line 2 (directly below, tighter line spacing): the vial size "
        f"'{vial_size}' in a lighter weight, smaller font, muted dark grey. "
        f"Both lines left-aligned with a small comfortable left margin. "
        f"Background of the filled area: same warm white as the rest of "
        f"the label (no off-color tinting, no shadows, no decorations)."
    )


def make_edit_mask() -> Path:
    """Build a 1024×1024 RGBA mask: opaque everywhere EXCEPT a rectangle
    covering the LEFT half of the label region. gpt-image-1's edit
    endpoint only modifies pixels where the mask is fully transparent
    (alpha=0), so the RUO badge, M watermark, cobalt dot, cap, glass,
    and background are all guaranteed pixel-identical.

    Coordinates calibrated to the base vial's label position (roughly
    x=360–700, y=585–815 of the 1024×1024 base). Editable zone is the
    left half of the label with a touch of padding."""
    from PIL import Image
    if MASK_PATH.exists():
        return MASK_PATH
    mask = Image.new("RGBA", (1024, 1024), (255, 255, 255, 255))  # all opaque
    # Editable rectangle — covers the left half of the label area where
    # compound text should land. Tuned to give the model enough room for
    # ~12 character compound names at a comfortable size.
    edit_box = Image.new("RGBA", (270, 240), (0, 0, 0, 0))  # transparent
    mask.paste(edit_box, (340, 580), edit_box)  # top-left corner of edit zone
    mask.save(MASK_PATH)
    return MASK_PATH


# ── Family classifier (mirrors lib/catalog-meta.ts familyByCompound) ───────
def family_for(compound: str) -> str:
    c = compound.lower()
    if re.search(r"(retatrutide|tirzepatide|semaglutide|cagrilintide|liraglutide)", c):
        return "glp1"
    if re.search(r"(selank|semax|pt-?141|melanotan|kisspeptin|dsip|oxytocin|vip|pe ?22)", c):
        return "neuropeptides"
    if re.search(r"(klow|glow|wolverine|\s\+\s)", c):
        return "blends"
    if re.search(r"(nad|ghk|mots|epitalon|amino|glutathione|5-amino|hcg|foxo)", c):
        return "cofactors"
    return "peptides"


def slugify(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[+/&]", "-", s)
    s = re.sub(r"[()]", "", s)
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")[:60]


def parse_inventory() -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(INVENTORY, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = [str(c) if c is not None else "" for c in r[:11]]
        sku = cells[0].strip()
        if not sku or sku.lower() == "sku": continue
        if "totals" in sku.lower() or "inventory position" in sku.lower(): continue
        product_name = cells[1].strip()
        vial_size = cells[2].strip()
        if not product_name or not vial_size: continue
        rows.append({"sku": sku, "compound": product_name, "vialSize": vial_size})
    return rows


def to_webp(src_png: Path, dest_webp: Path) -> bool:
    try:
        r = subprocess.run(
            ["cwebp", "-q", "85", "-quiet", str(src_png), "-o", str(dest_webp)],
            capture_output=True,
        )
        if r.returncode == 0:
            src_png.unlink(missing_ok=True)
            return True
    except FileNotFoundError:
        pass
    # Fallback: just rename .png → .webp's path with .png ext
    src_png.rename(dest_webp.with_suffix(".png"))
    return False


def gen_base(client, quality: str) -> Path:
    print(f"[BASE] generating canonical Merit vial → {BASE_PATH.relative_to(REPO_ROOT)}")
    result = client.images.generate(
        model="gpt-image-1",
        prompt=BASE_PROMPT,
        size="1024x1024",
        quality=quality,
        n=1,
    )
    b64 = result.data[0].b64_json
    if not b64:
        raise RuntimeError("gpt-image-1 returned no base image")
    BASE_PATH.write_bytes(base64.b64decode(b64))
    print(f"[BASE] ✓ saved ({BASE_PATH.stat().st_size // 1024}KB)")
    return BASE_PATH


def edit_for_sku(client, base_path: Path, compound: str, vial_size: str,
                 slug: str, quality: str) -> Path:
    """Single-shot gpt-image-1.generate per SKU with the full Merit label
    spec baked into the prompt. No flat overlay, no edit endpoint.
    Each SKU is its own AI render so the label conforms to the vial's
    curve and lighting naturally instead of looking pasted on.

    Cost: $0.04 (medium) or $0.19 (high) per SKU."""
    dest_png = OUTPUT_DIR / f"sku-{slug}.png"
    dest_webp = OUTPUT_DIR / f"sku-{slug}.webp"
    prompt = per_sku_prompt(compound, vial_size)
    result = client.images.generate(
        model="gpt-image-1",
        prompt=prompt,
        size="1024x1024",
        quality=quality,
        n=1,
    )
    b64 = result.data[0].b64_json
    if not b64:
        raise RuntimeError(f"gpt-image-1 returned no data for {slug}")
    dest_png.write_bytes(base64.b64decode(b64))
    to_webp(dest_png, dest_webp)
    return dest_webp if dest_webp.exists() else dest_png


def _legacy_pil_overlay(base_path: Path, compound: str, vial_size: str, slug: str) -> Path:
    """Retired PIL-overlay approach — flat label composited onto AI vial
    base. User feedback: looked pasted-on. Kept for reference only."""
    from PIL import Image, ImageDraw, ImageFont

    dest_png = OUTPUT_DIR / f"sku-{slug}.png"
    dest_webp = OUTPUT_DIR / f"sku-{slug}.webp"

    img = Image.open(base_path).convert("RGB").copy()

    # Use RGBA overlay so we can composite anti-aliased shapes with alpha
    overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    # ── Calibrated label region of the base vial ──────────────────────
    # Anchor is the label rectangle in the base image. Tuned by eye to
    # the typical position of the wrap-around label gpt-image-1 paints.
    LABEL_LEFT = 340
    LABEL_TOP = 535
    LABEL_WIDTH = 360
    LABEL_HEIGHT = 180

    # Brand colors
    COBALT = (46, 77, 219)     # #2E4DDB
    INK = (12, 16, 26)         # near-black
    SOFT = (110, 116, 130)     # muted dark grey for size line
    M_GREY = (216, 213, 205)   # #D8D5CD pale beige-grey for watermark
    PILL_TXT = (245, 245, 248) # near-white pill text

    HELVETICA = "/System/Library/Fonts/Helvetica.ttc"

    # ── (1) Large grey "M." watermark on right half ───────────────────
    # Sized to roughly match label height so it sits within the label
    # bounds without spilling onto the vial body. Drawn FIRST so other
    # elements can layer over it if needed.
    m_size = 150
    font_m = ImageFont.truetype(HELVETICA, size=m_size, index=1)  # Bold
    m_bbox = draw.textbbox((0, 0), "M", font=font_m)
    m_w = m_bbox[2] - m_bbox[0]
    m_h = m_bbox[3] - m_bbox[1]
    # Right edge of label minus dot+gap; vertical center of label
    m_x = LABEL_LEFT + LABEL_WIDTH - m_w - 50
    m_y = LABEL_TOP + (LABEL_HEIGHT - m_h) // 2 - 18
    draw.text((m_x, m_y), "M.", fill=(*M_GREY, 230), font=font_m)

    # ── (2) Cobalt dot to the right of the M ──────────────────────────
    dot_r = 9
    # Sit at the M's baseline (matches the period dot of M.)
    dot_cx = m_x + m_w + 26
    dot_cy = m_y + m_h - dot_r - 2
    draw.ellipse(
        [(dot_cx - dot_r, dot_cy - dot_r), (dot_cx + dot_r, dot_cy + dot_r)],
        fill=(*COBALT, 255),
    )

    # ── (3) RUO pill badge upper-left ─────────────────────────────────
    pill_text = "RESEARCH USE ONLY"
    pill_font_size = 11
    pill_font = ImageFont.truetype(HELVETICA, size=pill_font_size, index=1)
    pill_bbox = draw.textbbox((0, 0), pill_text, font=pill_font)
    pill_text_w = pill_bbox[2] - pill_bbox[0]
    pill_text_h = pill_bbox[3] - pill_bbox[1]
    pill_pad_x, pill_pad_y = 7, 3
    pill_w = pill_text_w + pill_pad_x * 2
    pill_h = pill_text_h + pill_pad_y * 2 + 4
    pill_x = LABEL_LEFT + 16
    pill_y = LABEL_TOP + 12
    # Rounded pill
    draw.rounded_rectangle(
        [(pill_x, pill_y), (pill_x + pill_w, pill_y + pill_h)],
        radius=pill_h // 2,
        fill=(*COBALT, 255),
    )
    draw.text(
        (pill_x + pill_pad_x, pill_y + pill_pad_y - 1),
        pill_text, fill=(*PILL_TXT, 255), font=pill_font,
    )

    # ── (4) Compound name (Helvetica Bold, auto-fit) ──────────────────
    # Fits inside the LEFT region of the label, ending before the M.
    title_size = 38
    font_bold = ImageFont.truetype(HELVETICA, size=title_size, index=1)
    bbox = draw.textbbox((0, 0), compound, font=font_bold)
    width = bbox[2] - bbox[0]
    # Max text width: left edge of label → m_x minus comfortable gap
    MAX_WIDTH = m_x - (LABEL_LEFT + 18) - 16
    while width > MAX_WIDTH and title_size > 16:
        title_size -= 2
        font_bold = ImageFont.truetype(HELVETICA, size=title_size, index=1)
        bbox = draw.textbbox((0, 0), compound, font=font_bold)
        width = bbox[2] - bbox[0]

    text_x = LABEL_LEFT + 18
    text_y = pill_y + pill_h + 8
    draw.text((text_x, text_y), compound, fill=(*INK, 255), font=font_bold)

    # ── (5) Vial size (Helvetica Regular, smaller, muted) ─────────────
    size_pt = max(14, int(title_size * 0.55))
    font_size = ImageFont.truetype(HELVETICA, size=size_pt, index=0)
    title_h = bbox[3] - bbox[1]
    draw.text(
        (text_x, text_y + title_h + 6),
        vial_size, fill=(*SOFT, 255), font=font_size,
    )

    # Composite the overlay onto the base
    final = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    final.save(dest_png, "PNG", optimize=True)
    to_webp(dest_png, dest_webp)
    return dest_webp if dest_webp.exists() else dest_png


def main():
    args = sys.argv[1:]
    base_only = "--base" in args
    do_all = "--all" in args
    example_slug = None
    if "--example" in args:
        i = args.index("--example")
        if i + 1 < len(args):
            example_slug = args[i + 1]
    only_filter = None
    if "--only" in args:
        i = args.index("--only")
        if i + 1 < len(args):
            only_filter = [s.strip().lower() for s in args[i + 1].split(",")]

    quality = "high" if ("--quality" in args and
                         args.index("--quality") + 1 < len(args) and
                         args[args.index("--quality") + 1] == "high") else "medium"

    if not (base_only or do_all or example_slug or only_filter):
        print("Pick a mode: --base | --example <slug> | --all | --only <a,b,c>")
        print("See module docstring for usage.")
        sys.exit(1)

    if not os.environ.get("OPENAI_API_KEY"):
        print("ERROR: OPENAI_API_KEY not set.")
        sys.exit(1)

    from openai import OpenAI
    client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])

    # NOTE: --base mode is now a no-op. Each SKU is a single-shot AI
    # render with the full Merit label spec baked into the prompt — no
    # shared base file is needed.
    if base_only:
        print("--base mode retired. Single-shot per-SKU prompts now bake the label spec directly. Use --example <slug> or --all.")
        return

    # Parse inventory
    rows = parse_inventory()
    plan: list[dict] = []
    for r in rows:
        slug = slugify(f"{r['compound']} {r['vialSize']}")
        plan.append({
            "sku": r["sku"],
            "compound": r["compound"],
            "vialSize": r["vialSize"],
            "slug": slug,
            "family": family_for(r["compound"]),
            "filename": f"sku-{slug}.webp",
            "imageUrl": f"/products/sku-{slug}.webp",
        })

    # Filter
    if example_slug:
        plan = [p for p in plan if p["slug"] == example_slug]
        if not plan:
            print(f"\nERROR: no inventory row matches slug '{example_slug}'.")
            print("Try one of:")
            for r in rows[:10]:
                hint = slugify(f"{r['compound']} {r['vialSize']}")
                print(f"  --example {hint}")
            sys.exit(1)
    elif only_filter:
        plan = [p for p in plan if any(f in p["compound"].lower() or f in p["slug"] for f in only_filter)]

    estimated_cost = len(plan) * (0.19 if quality == "high" else 0.04)
    print(f"\n[EDIT] plan: {len(plan)} SKU{'s' if len(plan) != 1 else ''}, quality={quality}, est ${estimated_cost:.2f}")

    if len(plan) > 5:
        print("Proceeding in 3 seconds. Ctrl-C to abort.")
        time.sleep(3)

    success = []
    failed = []
    for i, p in enumerate(plan, 1):
        prefix = f"[{i:>3}/{len(plan):>3}]"
        print(f"{prefix} → {p['compound']} {p['vialSize']}")
        try:
            out = edit_for_sku(client, BASE_PATH, p["compound"], p["vialSize"], p["slug"], quality)
            print(f"           ✓ {out.relative_to(REPO_ROOT)}")
            success.append(p)
        except Exception as e:
            print(f"           ✗ {type(e).__name__}: {e}")
            failed.append(p)

    # Update manifest with whatever we have so far (plus existing entries)
    full_plan = []
    for r in rows:
        slug = slugify(f"{r['compound']} {r['vialSize']}")
        full_plan.append({
            "sku": r["sku"],
            "compound": r["compound"],
            "vialSize": r["vialSize"],
            "slug": slug,
            "family": family_for(r["compound"]),
            "filename": f"sku-{slug}.webp",
            "imageUrl": f"/products/sku-{slug}.webp",
        })
    MANIFEST_PATH.write_text(json.dumps(full_plan, indent=2))

    print()
    print("═" * 60)
    print(f"Success: {len(success)} | Failed: {len(failed)}")
    if failed:
        print(f"Re-run failed ones with --only {','.join(p['slug'] for p in failed[:5])}")


if __name__ == "__main__":
    main()
