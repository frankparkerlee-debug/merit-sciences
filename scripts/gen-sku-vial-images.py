#!/usr/bin/env python3
"""
gen-sku-vial-images.py — Generate one branded vial image per SKU in the
Merit inventory sheet. Cap color encodes the product family so the
catalog grid reads visually grouped; the label carries the compound name
so each PDP has a unique-looking image.

USAGE
-----
  # Dry-run: list every row + planned file + family color
  python3 scripts/gen-sku-vial-images.py --dry-run

  # Generate (medium quality, ~$0.04/img → ~$2.68 for 67 SKUs)
  python3 scripts/gen-sku-vial-images.py

  # Skip rows whose output already exists (resume after a crash)
  python3 scripts/gen-sku-vial-images.py --skip-existing

  # Only generate a subset (handy for testing)
  python3 scripts/gen-sku-vial-images.py --only retatrutide,tirzepatide

  # Higher fidelity (~$0.19/img → ~$12.73 for 67)
  python3 scripts/gen-sku-vial-images.py --quality high

OUTPUT
------
  public/products/sku-<slug>.png            — one image per inventory row
  public/products/sku-image-manifest.json   — slug → {compound, vialSize, family} map
                                              read by the importer to set imageUrl

ENV
---
  OPENAI_API_KEY — pulled from ../Merit Peptides/.env
INVENTORY
---------
  ../Merit Peptides/Inventory 6.14.xlsx
"""

import os
import sys
import re
import json
import time
import base64
from pathlib import Path

HERE = Path(__file__).parent.parent
REPO_ROOT = HERE
PEPTIDES_ENV = REPO_ROOT.parent / "Merit Peptides" / ".env"
INVENTORY = REPO_ROOT.parent / "Merit Peptides" / "Inventory 6.14.xlsx"

# ── Load OPENAI_API_KEY ────────────────────────────────────────────────────
for env_path in (PEPTIDES_ENV, REPO_ROOT / ".env.local"):
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if "=" in line and not line.strip().startswith("#"):
                k, _, v = line.partition("=")
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

# ── Per-family cap color ───────────────────────────────────────────────────
# Pharmaceutical aluminum crimp cap colors chosen for brand cohesion +
# visual differentiation across the catalog grid. Matches the family
# classifier in lib/catalog-meta.ts (familyByCompound).
FAMILY_CAP = {
    "glp1":          ("deep saturated cobalt blue", "#2E4DDB"),
    "peptides":      ("dark gunmetal grey",         "#2A2D38"),
    "cofactors":     ("warm copper-amber metallic", "#B8702A"),
    "neuropeptides": ("deep royal purple",          "#5A3BB8"),
    "blends":        ("forest green",               "#2E5C3F"),
}


def family_for(compound: str) -> str:
    """Mirror of familyByCompound() in lib/catalog-meta.ts. Keep in sync."""
    c = compound.lower()
    if re.search(r"(retatrutide|tirzepatide|semaglutide|cagrilintide|liraglutide)", c):
        return "glp1"
    if re.search(r"(selank|semax|pt-?141|melanotan|kisspeptin|dsip|oxytocin|vip|pe ?22)", c):
        return "neuropeptides"
    # Blends require " + " (space-plus-space) so we don't catch "NAD+"
    # (a single compound name, not a combo).
    if re.search(r"(klow|glow|wolverine|\s\+\s)", c):
        return "blends"
    if re.search(r"(nad|ghk|mots|epitalon|amino|glutathione|5-amino|hcg|foxo)", c):
        return "cofactors"
    return "peptides"


def slugify(s: str) -> str:
    """Mirror of slugify() in inventory actions.ts."""
    s = s.lower()
    s = re.sub(r"[+/&]", "-", s)
    s = re.sub(r"[()]", "", s)
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")[:60]


def build_prompt(compound: str, vial_size: str, family: str) -> str:
    cap_desc, _ = FAMILY_CAP[family]
    short_name = compound.upper()
    # Keep on-label text minimal — gpt-image-1 mangles anything longer
    # than ~3 words on a small label. Sample short renderings:
    # "MERIT", "SCIENCES", compound name in single line.
    return (
        f"Photorealistic studio product photograph of a single pharmaceutical "
        f"peptide vial, centered, square 1:1 framing. The aluminum crimp cap "
        f"is {cap_desc}, slight metallic sheen, sharp specular highlight. "
        f"Clear glass body, lyophilized white powder visible at the bottom "
        f"(~one third of the vial). A clean minimalist WHITE label wrapped "
        f"around the lower half. The label reads: top line 'MERIT' in bold "
        f"sans-serif, small line 'SCIENCES' under it, and at the bottom a "
        f"single line '{short_name} · {vial_size}' in a smaller monospace "
        f"font. No other text anywhere. Soft warm studio lighting, gentle "
        f"shadow under the vial, shallow depth of field. The background is "
        f"a smooth warm cream off-white seamless studio sweep (#F4F1EA). No "
        f"props, no card, no documents, no humans, no hands. Apple-grade "
        f"editorial pharma photography. Absolutely photoreal — not an "
        f"illustration, not a 3D render."
    )


def parse_inventory() -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(INVENTORY, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = [str(c) if c is not None else "" for c in r[:11]]
        sku = cells[0].strip()
        if not sku or sku.lower() == "sku": continue
        if "totals" in sku.lower() or "inventory position" in sku.lower(): continue
        product_name = cells[1].strip()
        vial_size = cells[2].strip()
        if not product_name or not vial_size: continue
        rows.append({
            "sku": sku,
            "compound": product_name,
            "vialSize": vial_size,
        })
    return rows


def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    skip_existing = "--skip-existing" in args
    quality = "high" if ("--quality" in args and
                         args.index("--quality") + 1 < len(args) and
                         args[args.index("--quality") + 1] == "high") else "medium"
    only_filter = None
    if "--only" in args:
        idx = args.index("--only")
        if idx + 1 < len(args):
            only_filter = [s.strip().lower() for s in args[idx + 1].split(",")]

    if not INVENTORY.exists():
        print(f"ERROR: inventory not found at {INVENTORY}")
        sys.exit(1)

    rows = parse_inventory()
    print(f"Parsed {len(rows)} inventory rows from {INVENTORY.name}")

    output_dir = REPO_ROOT / "public" / "products"
    output_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = output_dir / "sku-image-manifest.json"

    # Plan: compute the slug + file path + prompt + family for each row.
    plan: list[dict] = []
    for r in rows:
        slug = slugify(f"{r['compound']} {r['vialSize']}")
        family = family_for(r['compound'])
        # Final filename is .webp — the script writes .png then converts
        # via cwebp at the end (called separately) to keep the repo small.
        # Manifest references .webp directly.
        filename = f"sku-{slug}.webp"
        plan.append({
            "sku": r["sku"],
            "compound": r["compound"],
            "vialSize": r["vialSize"],
            "slug": slug,
            "family": family,
            "filename": filename,
            "path": str((output_dir / filename).relative_to(REPO_ROOT)),
            "imageUrl": f"/products/{filename}",
        })

    if only_filter:
        plan = [p for p in plan if any(f in p["compound"].lower() or f in p["slug"] for f in only_filter)]
        print(f"Filtered to {len(plan)} rows matching: {only_filter}")

    print(f"\n{'SKU':<55} {'family':<14} {'file'}")
    print("─" * 110)
    for p in plan:
        print(f"  {p['sku'][:53]:<55} {p['family']:<14} {p['filename']}")

    if dry_run:
        print(f"\n[dry-run] {len(plan)} images would be generated. No API calls, no files written.")
        # Still write the manifest so the importer can resolve URLs against the planned layout
        manifest_path.write_text(json.dumps(plan, indent=2))
        print(f"[dry-run] Wrote planning manifest → {manifest_path.relative_to(REPO_ROOT)}")
        return

    if not os.environ.get("OPENAI_API_KEY"):
        print("\nERROR: OPENAI_API_KEY not set.")
        sys.exit(1)

    estimated_cost = len(plan) * (0.19 if quality == "high" else 0.04)
    print(f"\nQuality: {quality}. Estimated cost: ${estimated_cost:.2f}")
    print("Proceeding in 3 seconds. Ctrl-C to abort.")
    time.sleep(3)

    from openai import OpenAI
    client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])

    generated = 0
    skipped = 0
    failed = 0
    for i, p in enumerate(plan, 1):
        # gpt-image-1 returns PNG. We write to .png first, then convert
        # via cwebp at the end and remove the PNG. Final repo file is .webp.
        dest_png = output_dir / p["filename"].replace(".webp", ".png")
        dest_webp = output_dir / p["filename"]
        prefix = f"[{i:>3}/{len(plan):>3}]"
        if skip_existing and dest_webp.exists():
            print(f"{prefix} ⊝ {p['filename']} already exists, skipping")
            skipped += 1
            continue
        prompt = build_prompt(p["compound"], p["vialSize"], p["family"])
        print(f"{prefix} → {p['compound']} {p['vialSize']} ({p['family']})")
        try:
            result = client.images.generate(
                model="gpt-image-1",
                prompt=prompt,
                size="1024x1024",
                quality=quality,
                n=1,
            )
            b64 = result.data[0].b64_json
            if not b64:
                print(f"           ✗ no image data")
                failed += 1
                continue
            dest_png.write_bytes(base64.b64decode(b64))
            # Convert PNG → WebP at q=85 (≈85% size reduction, no visible
            # quality loss for product photography). Falls back to PNG if
            # cwebp isn't installed.
            import subprocess
            cwebp = subprocess.run(
                ["cwebp", "-q", "85", "-quiet", str(dest_png), "-o", str(dest_webp)],
                capture_output=True,
            )
            if cwebp.returncode == 0:
                dest_png.unlink()
                print(f"           ✓ {p['filename']}")
            else:
                # Keep the PNG, update the manifest entry to point at it
                p["filename"] = p["filename"].replace(".webp", ".png")
                p["imageUrl"] = p["imageUrl"].replace(".webp", ".png")
                print(f"           ✓ {p['filename']} (cwebp unavailable — saved as PNG)")
            generated += 1
        except Exception as e:
            print(f"           ✗ {type(e).__name__}: {e}")
            failed += 1

    # Always write the latest manifest, even partial
    manifest_path.write_text(json.dumps(plan, indent=2))

    print("\n" + "═" * 60)
    print(f"Generated:  {generated}")
    print(f"Skipped:    {skipped}")
    print(f"Failed:     {failed}")
    print(f"Manifest:   {manifest_path.relative_to(REPO_ROOT)}")
    print(f"Output:     {output_dir.relative_to(REPO_ROOT)}")
    if failed:
        print(f"\nRe-run with --skip-existing to retry only the failed rows.")


if __name__ == "__main__":
    main()
